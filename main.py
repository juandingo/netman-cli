import sys
import serial
import os
import re
import time
import logging
from collections import Counter
from openpyxl import load_workbook, Workbook
from serial_config import (
    load_config, save_config, list_available_ports,
    auto_detect_port, DEFAULT_CONFIG
)

config = load_config()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


def auto_detect_on_startup():
    global config
    port, ports_list, error = auto_detect_port()
    if port:
        config["port"] = port
        logger.info(f"Puerto detectado automáticamente: {port}")
        print(f"Puerto detectado automáticamente: {port}")
        save_config(config)
    elif error:
        print(f"\n{error}")
    elif ports_list:
        print("\nMúltiples puertos disponibles. Use la opción 5 para configurar.")
        for i, p in enumerate(ports_list, 1):
            print(f"  {i}. {p.device} - {p.description}")


def open_serial():
    if not config.get("port"):
        print("ERROR: No hay puerto seleccionado. Configure primero en opción 5.")
        return None
    try:
        ser = serial.Serial(
            port=config["port"],
            baudrate=config["baudrate"],
            bytesize=config["bits"],
            parity=config["parity"][0].upper() if config["parity"] else 'N',
            stopbits=config["stopbits"],
            timeout=config["timeout"],
            rtscts=False,
            dsrdtr=False
        )
        return ser
    except serial.SerialException as e:
        print(f"ERROR al conectar: {e}")
        if "PermissionError" in str(e) or "denied" in str(e).lower() or "13" in str(e):
            print()
            print("  ⚠︎  Windows: no se puede acceder al puerto", config["port"])
            print("  ─────────────────────────────────────────────")
            print("  1. Cierre otros programas que usen el puerto:")
            print("     PuTTY, Docklight, MobaXterm, HyperTerminal, etc.")
            print()
            print("  2. Abra el Administrador de Dispositivos y revise:")
            print("     - Que COM1 aparezca en 'Puertos (COM y LPT)'")
            print("     - Que no tenga un triángulo amarillo (error de driver)")
            print()
            print("  3. Ejecute esta terminal como ADMINISTRADOR:")
            print("     Clic derecho en CMD/PowerShell → 'Ejecutar como administrador'")
            print()
            print("  4. Para listar los puertos disponibles use la opción 5 → 1")
        return None


def send_command(ser, command, wait=0.5, total_timeout=15):
    ser.write((command + '\r\n').encode())
    output = b''
    last_data_time = None
    original_timeout = ser.timeout
    ser.timeout = 0.2
    deadline = time.time() + total_timeout
    try:
        while time.time() < deadline:
            data = ser.read(ser.in_waiting or 1)
            if data:
                output += data
                last_data_time = time.time()
                decoded = output.decode(errors='replace')
                if '<--- More --->' in decoded or '--More--' in decoded:
                    ser.write(b' ')
                    last_data_time = time.time()
            else:
                if last_data_time is not None and time.time() - last_data_time > wait:
                    break
    except:
        pass
    ser.timeout = original_timeout
    return output.decode(errors='replace')


def send_enable(ser):
    out = send_command(ser, "enable")
    if 'Password:' in out or 'password>' in out.lower():
        send_command(ser, "")
    return out


def flush_serial(ser):
    ser.reset_input_buffer()
    ser.reset_output_buffer()


def wake_up(ser):
    ser.write(b'\r')
    time.sleep(0.5)
    return read_all(ser)


def read_all(ser, timeout=2):
    output = b''
    last_data_time = None
    original_timeout = ser.timeout
    ser.timeout = 0.2
    deadline = time.time() + timeout
    try:
        while time.time() < deadline:
            data = ser.read(ser.in_waiting or 1)
            if data:
                output += data
                last_data_time = time.time()
            else:
                if last_data_time is not None and time.time() - last_data_time > 0.3:
                    break
    except:
        pass
    ser.timeout = original_timeout
    return output.decode(errors='replace')


def show_main_menu():
    print("\n" + "=" * 40)
    print("          MENÚ PRINCIPAL")
    print("=" * 40)
    print("1. Conectar al dispositivo")
    print("2. Enviar comando")
    print("3. Ver configuración actual")
    print("4. Información del dispositivo")
    print("5. Configuración serial")
    print("6. Exportar datos a Excel")
    print("7. Reiniciar a fábrica")
    print("8. Testear puertos del switch")
    print("9. Obtener info Allied Telesis FS750")
    print("0. Salir")
    print("=" * 40)


def cmd_connect():
    ser = open_serial()
    if ser:
        logger.info(f"Conectado a {config['port']} a {config['baudrate']} baud.")
        print(f"Conectado a {config['port']} a {config['baudrate']} baud.")
        ser.write(b'\r')
        time.sleep(1)
        respuesta = read_all(ser)
        if respuesta:
            for linea in respuesta.strip().split('\n'):
                linea = linea.strip()
                if linea:
                    print(f"  {linea}")
        else:
            print("AVISO: No se recibió respuesta del dispositivo.")
            print("  Verifique que el switch esté encendido y conectado al puerto serial.")
        ser.close()


def cmd_send_command():
    ser = open_serial()
    if not ser:
        return
    logger.info("Iniciando modo comando interactivo.")
    flush_serial(ser)
    init = wake_up(ser)
    if init:
        for l in init.strip().split('\n'):
            l = l.strip()
            if l:
                print(f"  {l}")
    else:
        print("AVISO: No se recibió respuesta del dispositivo.")
    print("\nModo comando interactivo. Escribe 'exit' para volver.\n")
    try:
        while True:
            cmd = input("Comando> ").strip()
            if cmd.lower() == 'exit':
                break
            if cmd:
                output = send_command(ser, cmd, wait=1)
                if output.strip():
                    print(output)
                else:
                    print("  (sin respuesta)")
    except KeyboardInterrupt:
        print("\nInterrumpido.")
    finally:
        ser.close()


def cmd_show_config():
    ser = open_serial()
    if not ser:
        return
    flush_serial(ser)
    wake_up(ser)
    out_enable = send_enable(ser)
    if '#' not in out_enable:
        logger.warning("No se detecta modo privilegiado (enable). La configuración puede estar incompleta.")
        print("  AVISO: No se detecta modo privilegiado. Verifique que no haya password de enable.")
    send_command(ser, "terminal length 0")
    send_command(ser, "pager lines 0")
    logger.info("Obteniendo running-config...")
    print("Obteniendo running-config...")
    output = send_command(ser, "show running-config", total_timeout=30)
    if output.strip():
        print(output)
    else:
        print("  No se recibió respuesta. Verifique la conexión con el switch/firewall.")
        print("  Asegúrese de estar en modo privilegiado (enable).")
    ser.close()


def cmd_device_info():
    ser = open_serial()
    if not ser:
        return
    flush_serial(ser)
    wake_up(ser)
    send_enable(ser)
    send_command(ser, "terminal length 0")
    send_command(ser, "pager lines 0")
    logger.info("Obteniendo información del dispositivo...")
    print("Obteniendo información del dispositivo...")
    output = send_command(ser, "show version", total_timeout=15)
    if output.strip():
        print(output)
    else:
        print("  No se recibió respuesta. Verifique la conexión con el switch/firewall.")
        print("  Asegúrese de que el dispositivo esté encendido y conectado al puerto serial.")
    ser.close()


def cmd_factory_reset():
    ser = open_serial()
    if not ser:
        return

    logger.info("Iniciando reinicio a fábrica.")
    try:
        flush_serial(ser)
        initial = wake_up(ser)

        if not initial.strip():
            print("AVISO: No se detecta respuesta del switch.")
            cont = input("  ¿Continuar de todas formas? (s/n): ").strip().lower()
            if cont != 's':
                print("Operación cancelada.")
                ser.close()
                return

        send_enable(ser)
        time.sleep(0.5)
        read_all(ser)

        print("\n" + "=" * 55)
        print("  ARCHIVOS EN LA FLASH DEL SWITCH")
        print("=" * 55)
        dir_out = send_command(ser, "dir flash:", wait=2)
        print(dir_out)

        print("=" * 55)
        print("  Archivos a eliminar (separados por coma):")
        print("  Ejemplo: vlan.dat, config.text")
        print("  'todo' para borrar todo (menos IOS)")
        print("  0 para saltar este paso")
        print("=" * 55)
        elegir = input("  ¿Qué archivos borrar?: ").strip().lower()

        borrar_lista = []
        if elegir == 'todo':
            archivos = re.findall(r'^\s*(\d+)\s+.+?([^\s]+\.\w+)', dir_out, re.MULTILINE)
            borrar_lista = [a for _, a in archivos if not a.endswith('.bin')]
        elif elegir != '0':
            borrar_lista = [a.strip() for a in elegir.split(',') if a.strip()]

        if borrar_lista:
            print(f"\n  Archivos seleccionados para borrar: {', '.join(borrar_lista)}")
            confirm = input("  ¿Confirmar borrado? (s/n): ").strip().lower()
            if confirm == 's':
                for archivo in borrar_lista:
                    print(f"  Borrando {archivo}...")
                    send_command(ser, f"delete flash:{archivo}")
                    time.sleep(0.5)
                    ser.write(b'\n')
                    time.sleep(0.5)
                    read_all(ser)
                print("  OK")
            else:
                print("  Borrado cancelado.")

        print("\n  Borrando startup-config...")
        send_command(ser, "write erase")
        time.sleep(1)
        ser.write(b'\n')
        time.sleep(1)
        read_all(ser)
        print("  OK")

        reiniciar = input("\n  ¿Reiniciar el switch ahora? (s/n): ").strip().lower()
        if reiniciar == 's':
            print("  Reiniciando...")
            send_command(ser, "reload")
            time.sleep(2)
            read_all(ser)
            ser.write(b'yes\n')
            time.sleep(1)
            read_all(ser)
            print("\n" + "!" * 50)
            print("   REINICIO COMPLETADO")
            print("   El switch se está reiniciando con valores de fábrica.")
            print("   Espere 2-3 minutos antes de intentar conectarse.")
            print("!" * 50)
        else:
            print("  No se reinició. Puede hacerlo manualmente con 'reload'.")

    except Exception as e:
        print(f"ERROR: {e}")
    finally:
        ser.close()


def cmd_export_excel():
    ser = open_serial()
    if not ser:
        return

    try:
        flush_serial(ser)
        wake_up(ser)

        logger.info("Conectando al switch para exportar datos...")
        print("Conectando al switch para obtener datos...")
        send_enable(ser)
        send_command(ser, "terminal length 0")
        send_command(ser, "pager lines 0")

        print("  Leyendo show version...")
        version_out = send_command(ser, "show version", total_timeout=15)
        if version_out.strip():
            print("  OK")
        else:
            print("  AVISO: No se recibió respuesta. Verifique la conexión.")

        print("  Leyendo show running-config...")
        running_out = send_command(ser, "show running-config", total_timeout=30)
        if running_out.strip():
            print("  OK")
        else:
            print("  AVISO: No se recibió respuesta.")

        device_type = detect_device_type(version_out) if version_out.strip() else "Switch"
        vlan_out = ""
        ip_int_out = ""
        serial_out = ""
        if device_type == "Switch":
            print("  Leyendo show vlan brief...")
            vlan_out = send_command(ser, "show vlan brief", total_timeout=10)
            if vlan_out.strip():
                print("  OK")
            else:
                print("  AVISO: No se recibió respuesta.")
        elif device_type == "Firewall":
            print("  Leyendo show interface ip brief...")
            ip_int_out = send_command(ser, "show interface ip brief", total_timeout=10)
            if ip_int_out.strip():
                print("  OK")
            else:
                print("  AVISO: No se pudo obtener interfaces IP.")
            print("  Leyendo show serial...")
            serial_out = send_command(ser, "show serial", total_timeout=10)
            if serial_out.strip():
                print("  OK")
            else:
                print("  AVISO: No se pudo obtener serial.")

        if not version_out.strip() and not running_out.strip():
            print(f"\nERROR: No se pudo obtener datos del {device_type.lower()}. Cancelando exportación.")
            return

        data = parse_device_data(version_out, running_out, vlan_out, serial_out)

        print("\n  Datos del dispositivo detectados. Complete los campos adicionales:")
        username = input("  Username: ").strip()
        data["Username"] = username
        password = input("  Password: ").strip()
        data["Password"] = password
        estado = input("  Estado (default: Activo): ").strip()
        data["Estado"] = estado if estado else "Activo"
        observaciones = input("  Observaciones: ").strip()
        data["Observaciones"] = observaciones

        filename = input("\nNombre archivo Excel (default: dispositivos.xlsx): ").strip()
        if not filename:
            filename = "dispositivos.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        append_excel_row(data, filename)
        logger.info(f"Datos exportados a: {filename}")
        print(f"\n  Datos exportados a: {filename}")
        print("  Resumen de datos extraídos:")
        for k, v in data.items():
            if v:
                print(f"    {k}: {v}")
            else:
                print(f"    {k}: (no detectado)")

    except Exception as e:
        logger.error(f"Error al obtener datos: {e}")
        print(f"ERROR al obtener datos: {e}")
    finally:
        ser.close()


def detect_brand(version_text):
    if re.search(r'cisco', version_text, re.IGNORECASE):
        return "Cisco"
    if re.search(r'Hewlett.?Packard|HP\s+ProCurve|ProCurve', version_text, re.IGNORECASE):
        return "HP"
    if re.search(r'Dell.*(?:PowerConnect|Force10|Networking)', version_text, re.IGNORECASE):
        return "Dell"
    if re.search(r'Juniper', version_text, re.IGNORECASE):
        return "Juniper"
    if re.search(r'Extreme', version_text, re.IGNORECASE):
        return "Extreme"
    if re.search(r'Brocade', version_text, re.IGNORECASE):
        return "Brocade"
    if re.search(r'MikroTik|RouterOS', version_text, re.IGNORECASE):
        return "MikroTik"
    if re.search(r'Netgear', version_text, re.IGNORECASE):
        return "Netgear"
    if re.search(r'TP.?Link', version_text, re.IGNORECASE):
        return "TP-Link"
    return "Switch genérico"


def detect_device_type(version_text):
    if re.search(r'Adaptive Security Appliance|Security Appliance|PIX\s|ASA\s?\d+|Palo\s*Alto|PA-\s*\d+|FortiGate|Fortinet|pfSense|OPNsense|SonicWALL|SonicOS|Check\s*Point|Firewall\s+(Module|Appliance)', version_text, re.IGNORECASE):
        return "Firewall"
    return "Switch"


def parse_interfaces(version_text):
    interfaces = []
    for m in re.finditer(r'(\d+)\s+(.+?)\s+[Ii]nterface', version_text):
        count = m.group(1)
        tipo = m.group(2).strip().rstrip('/,;')
        if tipo:
            interfaces.append(f"{count} {tipo}")
    if not interfaces:
        for m in re.finditer(r'(\d+)\s+(FastEthernet|Gigabit Ethernet|GigabitEthernet|TenGigabitEthernet|Ports)', version_text, re.IGNORECASE):
            count, tipo = m.group(1), m.group(2)
            if tipo:
                interfaces.append(f"{count} {tipo}")
    if not interfaces:
        ifaces = re.findall(r'\d+:\s+\w+:\s+(\S+)\s*:', version_text)
        if ifaces:
            interfaces = list(dict.fromkeys(ifaces))
    return ", ".join(interfaces) if interfaces else ""


def parse_device_data(version_text, running_text, vlan_text="", serial_text=""):
    data = {
        "Tipo": detect_device_type(version_text),
        "Marca": detect_brand(version_text),
        "Modelo": "",
        "Revisión": "",
        "Firmware": "",
        "Número de serie": "",
        "Hostname": "",
        "MAC": "",
        "IP-gestión": "",
        "Username": "",
        "Password": "",
        "Licencia": "",
        "Estado": "",
        "Observaciones": ""
    }

    # Modelo
    m = re.search(r'Model\s+number\s*[=:]\s*(\S+)', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'((?:WS-C|CGS|IE-\d+|SG\d+|CSR)\S+)', version_text)
    if not m:
        m = re.search(r'cisco\s+(\S+)\s+\(', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'[Pp]roduct\s+(?:name|number|model)\s*[=:]\s*(\S+)', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'[Mm]odel\s*(?:name|number|No\.?)?\s*[=:.]?\s*(\S+)', version_text)
    if not m:
        m = re.search(r'Hardware:\s*([^,\s]+)', version_text)
    if m:
        data["Modelo"] = m.group(1).strip().rstrip(',')

    # Revisión del modelo
    m = re.search(r'Model\s+revision\s+number\s*[=:]\s*(\S+)', version_text, re.IGNORECASE)
    if m:
        data["Revisión"] = m.group(1).strip()
    else:
        m = re.search(r'[Mm]odel\s+rev\S*\s*[=:.]?\s*(\S+)', version_text)
        if m:
            data["Revisión"] = m.group(1).strip()
    if not data["Revisión"]:
        m = re.search(r'\(revision\s+(\S+)\)', version_text, re.IGNORECASE)
        if m:
            data["Revisión"] = m.group(1).strip()

    # Firmware
    m = re.search(r'Version\s+(\S[\w().]*)', version_text)
    if not m:
        m = re.search(r'[Ff]irmware\s*(?:version|rev)?\s*[=:]\s*(\S+)', version_text)
    if not m:
        m = re.search(r'Software\s+(\d+\.\d+\(?\d*\)?[^,\s]*)', version_text)
    if m:
        data["Firmware"] = m.group(1).strip().rstrip(',')

    # Número de serie
    m = re.search(r'^Syste?m?\s+[Ss]erial\s+[Nn]umber\s*[=:]\s*(\S+)', version_text, re.MULTILINE)
    if not m:
        m = re.search(r'[Pp]rocessor\s+board\s+ID\s+(\S+)', version_text)
    if not m:
        m = re.search(r'[Bb]oard\s+ID\s+(\S+)', version_text)
    if m:
        data["Número de serie"] = m.group(1).strip()
    if not data["Número de serie"] and 'Motherboard' in version_text:
        for line in version_text.split('\n'):
            if 'serial number' in line.lower() and 'motherboard' not in line.lower() and 'power supply' not in line.lower():
                parts = line.split(':')
                if len(parts) >= 2:
                    data["Número de serie"] = parts[-1].strip()
                    break
    if not data["Número de serie"]:
        m = re.search(r'[Ss][Nn][:=]\s*(\S+)', version_text)
        if m:
            data["Número de serie"] = m.group(1).strip()
    if not data["Número de serie"]:
        m = re.search(r'[Ss]erial\s+[Nn]umber\s*:?\s*(\S+)', version_text)
        if m:
            data["Número de serie"] = m.group(1).strip()
    if not data["Número de serie"] and serial_text:
        m = re.search(r'(\S+)', serial_text)
        if m:
            data["Número de serie"] = m.group(1).strip()

    # MAC
    m = re.search(r'(?:Base\s+)?(?:ethernet\s+)?MAC\s+(?:Address|Router)\s*[=:]\s*(\S+)', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'([0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4})', version_text)
    if not m:
        m = re.search(r'([0-9A-Fa-f]{2}[:-][0-9A-Fa-f]{2}[:-][0-9A-Fa-f]{2}[:-][0-9A-Fa-f]{2}[:-][0-9A-Fa-f]{2}[:-][0-9A-Fa-f]{2})', version_text)
    if m:
        data["MAC"] = m.group(1).strip()

    # Hostname
    m = re.search(r'hostname\s+(\S+)', running_text)
    if not m and version_text:
        m = re.search(r'^(\S+)\s+up\s+\d+\s+(sec|min|hour|day|week)', version_text, re.MULTILINE)
    if m:
        data["Hostname"] = m.group(1).strip()

    # IP gestión
    m = re.search(r'interface\s+Vlan1.*?ip\s+address\s+(\S+)\s+(\S+)', running_text, re.DOTALL)
    if not m:
        m = re.search(r'ip\s+address\s+(\S+)\s+\S+', running_text)
    if not m:
        m = re.search(r'interface\s+vlan.*?ip\s+address\s+(\S+)', running_text, re.DOTALL | re.IGNORECASE)
    if m:
        data["IP-gestión"] = m.group(1).strip()

    # Licencia
    m = re.search(r'(License\s*[^:\n]*[:]\s*.+)', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'(image\s+license\s*[:-]\s*.+)', version_text, re.IGNORECASE)
    if not m:
        m = re.search(r'(Running\s+\w+\s+Image)', version_text, re.IGNORECASE)
    if m:
        data["Licencia"] = m.group(1).strip()
    else:
        data["Licencia"] = "N/A"

    # Interfaces
    interfaces = parse_interfaces(version_text)
    if not interfaces:
        int_running = re.findall(r'^interface\s+(FastEthernet|GigabitEthernet|TenGigabitEthernet|Port-Channel|Loopback)\S*', running_text, re.MULTILINE)
        if int_running:
            from collections import Counter
            counts = Counter(int_running)
            interfaces = ", ".join(f"{v} {k}" for k, v in counts.items())
    data["Interfaces"] = interfaces

    return data


def parse_allied_telesis(output):
    data = {
        "Tipo": "Switch",
        "Marca": "Allied Telesis",
        "Modelo": "",
        "Puertos": "",
        "Firmware": "",
        "Boot Loader": "",
        "Número de serie": "",
        "Hostname": "",
        "MAC": "",
        "IP-gestión": "",
        "Username": "manager",
        "Password": "friend",
        "Licencia": "N/A",
        "Estado": "",
        "Observaciones": ""
    }

    m = re.search(r'(AT-FS\S+)', output)
    if m:
        data["Modelo"] = m.group(1)

    m = re.search(r'Runtime Image\s*:\s*Version\s+(.+)', output)
    if m:
        data["Firmware"] = m.group(1).strip()

    m = re.search(r'Boot Loader\s*:\s*Version\s+(.+)', output)
    if m:
        data["Boot Loader"] = m.group(1).strip()

    m = re.search(r'Switch Name\s*:\s*(.*)', output)
    if m:
        data["Hostname"] = m.group(1).strip()

    m = re.search(r'MAC Address\s*:\s*(\S+)', output)
    if m:
        data["MAC"] = m.group(1)

    m = re.search(r'IP Address\s*:\s*(\S+)', output)
    if m:
        data["IP-gestión"] = m.group(1)

    if data["Modelo"]:
        m = re.search(r'/(\d+)', data["Modelo"])
        if m:
            data["Puertos"] = m.group(1)

    return data


FIELDNAMES = [
    "ID", "Tipo", "Marca", "Modelo", "Puertos", "Firmware", "Boot Loader",
    "Número de serie", "Hostname", "MAC", "IP-gestión", "Username", "Password",
    "Licencia", "Estado", "Observaciones"
]


def append_excel_row(data, filename):
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        ws = wb.active
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                try:
                    max_id = max(max_id, int(row[0]))
                except (ValueError, TypeError):
                    pass
        row_id = max_id + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(FIELDNAMES)
        for col, ancho in zip('ABCDEFGHIJKLMNOP', [6, 10, 14, 20, 8, 16, 16, 18, 16, 20, 18, 14, 14, 22, 12, 30]):
            ws.column_dimensions[col].width = ancho
        row_id = 1
    row_data = [row_id] + [data.get(f, '') for f in FIELDNAMES[1:]]
    ws.append(row_data)
    wb.save(filename)


def parse_port_status(output):
    puertos = []
    encabezado = True
    for linea in output.strip().split('\n'):
        linea = linea.strip()
        if not linea or linea.startswith('--'):
            continue
        if encabezado:
            if 'Port' in linea or 'Interface' in linea:
                encabezado = False
                continue
        partes = linea.split()
        if len(partes) >= 3:
            nombre = partes[0]
            if re.match(r'[A-Za-z]+[\d/]+', nombre):
                estado = partes[2].lower()
                puertos.append((nombre, estado))
    return puertos


def show_port_status(ser):
    output = send_command(ser, "show interfaces status", wait=2)
    if not output.strip():
        output = send_command(ser, "show interfaces description", wait=2)
    if not output.strip():
        print("  No se pudo obtener el estado de los puertos.")
        return []

    puertos = parse_port_status(output)
    if not puertos:
        print("  No se encontraron puertos en la salida del comando.")
        return []

    total = len(puertos)
    up = sum(1 for _, e in puertos if e == 'connected' or e == 'up')
    down = sum(1 for _, e in puertos if e == 'notconnect' or e in ('down', 'disabled'))

    print(f"\n{'='*60}")
    print(f"  PUERTOS DEL SWITCH ({total} detectados)")
    print(f"{'='*60}")
    for nombre, estado in puertos:
        icono = "🟢" if estado in ('connected', 'up') else "🔴"
        print(f"  {icono} {nombre:12} {estado}")
    print(f"{'='*60}")
    print(f"  Total: {total}  |  Conectados: {up}  |  Desconectados: {total - up}")
    print(f"{'='*60}")
    return puertos


def check_port_linked(ser, nombre, timeout=5):
    inicio = time.time()
    while time.time() - inicio < timeout:
        output = send_command(ser, "show interfaces status", wait=0.5)
        if not output.strip():
            output = send_command(ser, "show interfaces description", wait=0.5)
        for puerto, est in parse_port_status(output):
            if puerto == nombre and est in ('connected', 'up'):
                return True
        time.sleep(1)
    return False


def interactive_port_test(ser, puertos):
    down_ports = [p for p in puertos if p[1] not in ('connected', 'up')]
    if not down_ports:
        print("\n  Todos los puertos están conectados. No hay nada que testear.")
        return

    print(f"\n  {'='*60}")
    print(f"  MODO TEST MANUAL")
    print(f"  {'='*60}")
    print(f"  Puertos a testear: {len(down_ports)}")
    print(f"  Conecte un cable de red desde un equipo activo")
    print(f"  al puerto indicado. El programa detectará cuando")
    print(f"  el enlace se levante (timeout: 5s por puerto).")
    print(f"  {'='*60}\n")

    for i, (nombre, estado_actual) in enumerate(down_ports, 1):
        print(f"  [{i}/{len(down_ports)}] Conecte un cable al puerto {nombre}")
        input("  Presione Enter cuando esté listo...")

        send_enable(ser)
        time.sleep(0.3)
        read_all(ser)

        transcurrido = time.time()
        ok = check_port_linked(ser, nombre, timeout=5)
        transcurrido = time.time() - transcurrido

        if ok:
            print(f"  ✅ Puerto {nombre} — CONECTADO ({transcurrido:.0f}s)")
        else:
            print(f"  ❌ Puerto {nombre} — SIN RESPUESTA (5s agotados)")

    print(f"\n  {'='*60}")
    print(f"  PRUEBA COMPLETADA")
    print(f"  {'='*60}")


def cmd_test_ports():
    ser = open_serial()
    if not ser:
        return
    logger.info("Iniciando test de puertos.")
    try:
        flush_serial(ser)
        wake_up(ser)
        send_enable(ser)
        time.sleep(0.5)
        read_all(ser)

        puertos = show_port_status(ser)
        if not puertos:
            ser.close()
            return

        op = input("\n  ¿Ejecutar test manual de puertos? (s/n): ").strip().lower()
        if op == 's':
            interactive_port_test(ser, puertos)

    except Exception as e:
        print(f"ERROR: {e}")
    finally:
        ser.close()


def read_until(ser, target, timeout=15):
    output = ""
    original_timeout = ser.timeout
    ser.timeout = 0.2
    deadline = time.time() + timeout
    last_data_time = time.time()
    try:
        while time.time() < deadline:
            try:
                data = ser.read(ser.in_waiting or 1)
            except:
                break
            if data:
                output += data.decode(errors='replace')
                last_data_time = time.time()
                if target in output:
                    return output
            else:
                if time.time() - last_data_time > 2:
                    break
    finally:
        ser.timeout = original_timeout
    return output


def detect_allied_telesis_state(output):
    if "Login:" in output or "login:" in output:
        return "login"
    if "Main Menu" in output and "Command>" in output:
        return "menu"
    if "General Information" in output:
        return "info"
    return "unknown"


def cmd_allied_telesis():
    ser = open_serial()
    if not ser:
        return

    try:
        flush_serial(ser)
        ser.reset_input_buffer()

        print("  Conectando al Allied Telesis FS750...")

        ser.write(b'\r')
        time.sleep(0.5)
        output = read_all(ser, timeout=3)

        if not output.strip():
            ser.write(b'\r')
            time.sleep(0.5)
            output = read_all(ser, timeout=3)

        state = detect_allied_telesis_state(output)

        if state == "login":
            print("  Login detectado. Enviando credenciales manager/friend...")
            send_command(ser, "manager", wait=0.5, total_timeout=5)
            time.sleep(0.3)
            output = send_command(ser, "friend", wait=1.5, total_timeout=10)
            if 'Main Menu' not in output and 'Command>' not in output:
                print("  Error: No se pudo iniciar sesión.")
                print(output[-400:])
                return
            print("  Sesión iniciada correctamente.")
        elif state == "menu":
            print("  Ya en el menú principal. Saltando login...")
        elif state == "info":
            print("  Ya en General Information. Parseando directamente...")
            data = parse_allied_telesis(output)
            print(output)
        else:
            print("  No se reconoce el estado del switch.")
            print(output[-500:])
            return

        if state != "info":
            ser.reset_input_buffer()
            time.sleep(0.3)
            print("  Obteniendo General Information...")
            output = send_command(ser, "G", wait=2, total_timeout=15)
            if not output.strip():
                print("  AVISO: No se recibió respuesta para General Information.")
                return
            print(output)
            data = parse_allied_telesis(output)

        print("\n  Datos detectados del Allied Telesis:")
        for k, v in data.items():
            if v:
                print(f"    {k}: {v}")
            else:
                print(f"    {k}: (no detectado)")

        print("\n  Complete los campos adicionales:")
        serial_num = input("  Número de serie (deje vacío si no disponible): ").strip()
        if serial_num:
            data["Número de serie"] = serial_num
        estado = input("  Estado (default: Activo): ").strip()
        data["Estado"] = estado if estado else "Activo"
        observaciones = input("  Observaciones: ").strip()
        data["Observaciones"] = observaciones

        filename = input("\n  Nombre archivo Excel (default: dispositivos.xlsx): ").strip()
        if not filename:
            filename = "dispositivos.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        append_excel_row(data, filename)
        logger.info(f"Datos exportados a: {filename}")
        print(f"\n  Datos exportados a: {filename}")

    except Exception as e:
        logger.error(f"Error en Allied Telesis: {e}")
        print(f"ERROR: {e}")
    finally:
        ser.close()


def show_serial_menu():
    while True:
        print("\n" + "-" * 35)
        print("    CONFIGURACIÓN SERIAL")
        print("-" * 35)
        print(f"  Puerto:   {config.get('port', 'No seleccionado')}")
        print(f"  Baudrate: {config.get('baudrate')}")
        print(f"  Bits:     {config.get('bits')}")
        print(f"  Paridad:  {config.get('parity')}")
        print(f"  Stop bits:{config.get('stopbits')}")
        print(f"  Timeout:  {config.get('timeout')}s")
        print("-" * 35)
        print("1. Listar puertos disponibles")
        print("2. Seleccionar puerto")
        print("3. Configurar baudrate")
        print("4. Probar conexión")
        print("0. Volver")
        print("-" * 35)

        choice = input("Opción: ").strip()

        if choice == "1":
            ports = list_available_ports()
            if not ports:
                print("No se encontraron puertos seriales.")
            else:
                for i, p in enumerate(ports, 1):
                    print(f"  {i}. {p.device} - {p.description}")

        elif choice == "2":
            ports = list_available_ports()
            if not ports:
                print("No se encontraron puertos seriales.")
                continue
            print("Puertos disponibles:")
            for i, p in enumerate(ports, 1):
                print(f"  {i}. {p.device} - {p.description}")
            try:
                sel = int(input("Seleccione: ")) - 1
                if 0 <= sel < len(ports):
                    config["port"] = ports[sel].device
                    save_config(config)
                else:
                    print("Selección inválida.")
            except ValueError:
                print("Entrada inválida.")

        elif choice == "3":
            try:
                baud = int(input("Baudrate (ej: 9600, 19200, 115200): "))
                config["baudrate"] = baud
                save_config(config)
            except ValueError:
                print("Valor inválido.")

        elif choice == "4":
            if not config.get("port"):
                print("No hay puerto seleccionado.")
                continue
            try:
                ser = serial.Serial(
                    port=config["port"],
                    baudrate=config["baudrate"],
                    bytesize=config["bits"],
                    parity=config["parity"][0].upper() if config["parity"] else 'N',
                    stopbits=config["stopbits"],
                    timeout=config["timeout"]
                )
                print(f"Conexión exitosa a {config['port']}")
                ser.close()
            except serial.SerialException as e:
                print(f"ERROR: {e}")

        elif choice == "0":
            break


def main():
    global config
    config = load_config()

    if not config.get("port"):
        auto_detect_on_startup()

    while True:
        show_main_menu()
        choice = input("Opción: ").strip()

        if choice == "1":
            cmd_connect()
        elif choice == "2":
            cmd_send_command()
        elif choice == "3":
            cmd_show_config()
        elif choice == "4":
            cmd_device_info()
        elif choice == "5":
            show_serial_menu()
        elif choice == "6":
            cmd_export_excel()
        elif choice == "7":
            cmd_factory_reset()
        elif choice == "8":
            cmd_test_ports()
        elif choice == "9":
            cmd_allied_telesis()
        elif choice == "0":
            print("Saliendo...")
            sys.exit(0)
        else:
            print("Opción inválida.")


if __name__ == "__main__":
    main()
